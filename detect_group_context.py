import os
from dotenv import load_dotenv
import numpy as np
import pandas as pd
from pandasai import SmartDataframe
from pandasai import SmartDatalake
from pandasai import Agent
# Instantiate a LLM
from pandasai.llm import OpenAI
import re
import warnings
import logging
from openpyxl import load_workbook
from qworks_logger import logger
from config.default import Default
import generic_context as generic_context

load_dotenv()

# os.environ["OPENAI_API_KEY"] = ""
# os.environ["PANDASAI_API_KEY"] = ""
os.environ.pop("PANDASAI_API_KEY", None)

# Filter ignore warning log from openpyxl module
# E.g: UserWarning: Data Validation extension is not supported and will be removed
warnings.simplefilter('ignore', category=UserWarning)

### QWorks code
llm = OpenAI(
    temperature = 0.5,
    seed = 26
)
# Replace file path with with the actual path to your Excel file.
file_path = Default.DATASET_FILE_PATH

CONFIG = {
    "llm": llm, 
    "enable_cache": False, 
    "enforce_privacy": True,
    "max_retries": 10
}

def client(text, role="assistant"):
    completion = llm.client.create(
            model="gpt-3.5-turbo", # gpt-3.5-turbo-instruct
            messages=[
                {"role": role, "content": text}
            ]
        )
    return completion.choices[0].message.content

def get_sub_df(sheet_name, headers):
    columns = []
    index = None
    old_col_to_new_col_mapping = {}

    sub_df = pd.read_excel(file_path, sheet_name=sheet_name)

    for header in headers:
        row, col = np.where(sub_df == header)
        index = (sub_df.index[row])[0]
        column = (sub_df.columns[col])[0]
        columns.append(column)
        old_col_to_new_col_mapping[column] = header

    filtered_df = sub_df.query('index > @index')[columns]
#     filtered_df = filtered_df.dropna()

    data_frame = filtered_df.rename(columns=old_col_to_new_col_mapping)
    return data_frame

def init_lake(dfs):
    # Initialize SmartDataLake object
    sdl = SmartDatalake(dfs, config=CONFIG)
    return sdl

def init_smart_data_frame(sub_data_frame, sheet_name):
    # Initialize SmartDataframe object
    sdf = SmartDataframe(sub_data_frame, name=sheet_name, description='The ' + sheet_name +' of an Offer Case template', config=CONFIG)
    return sdf

def check_match(string):
    result = None
    pattern = r'\[GROUP_(\d+)\]'
    match = re.search(pattern, string)
    if match:
        result = match.group(1)  # Extract the value within square brackets
    
    return result
    

def suggested_replies_default():
    suggested = []
    suggested.append("What kind of template is this?")
    suggested.append("What action should be taken on this file?")
    suggested.append("Which customer is this template representing?")
    suggested.append("What is the location of the customer?")
    suggested.append("How many project phases are there?")
    suggested.append("What is the List of Project phases and their respective start dates?")
    suggested.append("How many tabs/worksheets are there in the file and how are they related?")
    suggested.append("How many currencies are involved in the offer/contract transaction?")
    suggested.append("What are the products listed in the BoQ (Bill of Quantities) tab?")
    suggested.append("Are there any duplicate entries within each individual sheet?")
    suggested.append("Show all pricing schemes?")
    suggested.append("What is the pricing discount across different pricing schemes?")
    suggested.append("Counts of Type 1 & 2 CPBs in the dataset?")
    
    return suggested
    
def suggested_replies(group, user_input):
    suggested = []
    if not group:
        suggested = suggested_replies_default()
    else:
        question_group = question_by_group()
        suggested = question_group[group]
    
    if user_input in suggested:
        suggested.remove(user_input)
    
    return suggested

# Create a dictionary to map input values to question group
def question_by_group():
    question_by_group = {
        "1": ["What kind of template is this?", "What type of template is uploaded?", "What is the Case Type of this Excel file?", "What is this?"],
        "2": ["What action should be taken on this file?", "What is the upload type?", "What steps are required for this file?", "Could you specify the upload type for this document?", "How should this file be processed?", "Can you identify the upload type of this document?", "What categorization does this file have in terms of upload?", "What upload type is associated with this file?", "What upload classification does this document have?", "What action is needed for this file\'s upload type?"],
        "3": ["Which customer is this template representing?", "Who is the customer?", "For whom is this template designed, in terms of the end customer?", "Who is the ultimate recipient or end user of this template?", "Which party or entity serves as the end customer for this template?", "Who is the target end customer addressed by this template?", "Whose needs or requirements does this template cater to as the end customer?", "Who is the intended final user of this template?", "Who is the primary customer identified in this template?", "Who is the ultimate consumer or beneficiary of this template?", "To whom does this template aim to provide value as the end customer?", "Who is the end recipient of the services/products outlined in this template?"],
        "4": ["What is the location of the customer?", "Where is the customer located?", "In terms of Location Management, what country is the customer associated with?", "Which country is specified in the Location Management for the customer?", "Regarding Location Management, what is the country of the customer?", "In the context of Location Management, which country does the customer belong to?", "What country name is stored under Location Management for the customer?", "Concerning Location Management, which country is attributed to the customer?", "Under Location Management, what country is linked to the customer\'s profile?", "When considering Location Management, what country is mentioned for the customer?", "In the perspective of Location Management, what is the country assigned to the customer?", "Within Location Management, which country is associated with the customer\'s details?"],
        "5": ["How many project phases are there?", "List the project phases?", "What is the List of Project phases?", "Can you provide me with the List of Project phases?","List down the Project phases, please.","What are the phases involved in the project?", "I need the Project phases in a list format, please.","Could you please outline the Project phases?","How many Project phases are there?", "Enumerate the phases of the project.","What does the sequence of Project phases look like?","Describe the Project phases, please."],
        "6": ["What is the List of Project phases and their respective start dates?","Can you provide me with the List of Project phases along with the start date of Phase [number]?","List down the Project phases, and could you also share the start date of Phase [number]?","What are the phases involved in the project, and what\'s the start date of Phase [number]?","I need the Project phases in a list format, along with the start date of Phase [number], please.","Could you please outline the Project phases, including the start date of Phase [number]?","How many Project phases are there, and when does Phase [number] start?","Enumerate the phases of the project and provide the start date of Phase [number].","What does the sequence of Project phases look like, and when does Phase [number] begin?","Describe the Project phases, please, and also share the start date of Phase [number]"],
        "7": ["How many tabs/worksheets are there in the file and how are they related?","How are the tabs/worksheet in this Excel file related to each other?","How do the tabs/worksheet in this Excel file interrelate with each other?","How are the tabs or worksheets structured in the Excel file?","What information is contained within each tab or worksheet?","Are there any common themes or categories among the tabs?","How do the tabs or worksheets interact with each other?","Are there any dependencies between the tabs or worksheets?","Do certain tabs or worksheets serve as inputs for others?","Are there any calculations or analyses that span multiple tabs?","How does the data flow between different tabs or worksheets?","Are there any key identifiers or references used across multiple tabs?","Can you describe the purpose of each tab or worksheet?","Are there any relationships between the tabs in terms of project phases?","Do the tabs represent different stages of a project or process?","How do the tabs contribute to the overall structure of the offer case template?","Are there specific tabs dedicated to certain aspects such as products or project phases?","What role does the data dictionary tab play in relation to other tabs?"],
        "8": ["How many currencies are involved in the offer/contract transaction?","What are the primary currencies in the transaction?", "Can you provide a breakdown of the currencies exchanged in the offer/contract?","What is the base currency used in the transaction?", "Are there any secondary currencies involved, and if so, what are they?","What is the exchange rate for each currency pair in the offer/contract?", "Is there a preferred currency for settlement, and if yes, what is it?","Are there any currency conversion fees associated with the transaction?", "How does the exchange rate affect the overall value of the offer/contract?","Can you specify the currency denominations used in the transaction?"],
        "9": ["What are the products listed in the BoQ (Bill of Quantities) tab?","How many items are included in the BoQ?", "What are the quantities specified for each product in the BoQ?","What are the unit prices listed for each product in the BoQ?", "Can you provide a breakdown of the total cost for each product in the BoQ?","Are there any specific notes or comments associated with products in the BoQ?", "Is there any additional information related to products in the BoQ that needs to be considered?","Are there any discrepancies or inconsistencies in the data within the BoQ?", "Are there any missing entries or incomplete data in the BoQ?","Can you confirm the accuracy of the data provided in the BoQ?","Are all the contextual data used with the product BoQ?", "Is there any contextual data present that does not correspond to any product listed in the BoQ?","Can you identify any contextual data that was not utilized or referenced within the BoQ?", "What is the extent of unused contextual data in relation to the products listed in the BoQ?","Are there any specific patterns or reasons why certain contextual data was not incorporated into the BoQ?"],
        "10": ["Are there any duplicate entries within each individual sheet?","How many duplicate entries are there in each sheet?","Can you identify which columns have duplicates within each sheet?","Are there any duplicate entries that span across multiple sheets?","Which specific rows contain duplicate data within each sheet?","Can you provide a summary of the total number of duplicates across all sheets?","Do any sheets have a particularly high number of duplicate entries compared to others?","Are there any patterns or trends in the types of data that are duplicated?","Can you identify any potential causes for the presence of duplicates within the data?","How would you recommend addressing the issue of duplicates within the Excel file?"],
        "11": ["Show all pricing schemes?","List all pricing schemes","What are the available pricing schemes?","Show me a list of all pricing schemes.", "Can you display the pricing schemes currently active?","List all pricing schemes along with their details.","How many pricing schemes are there in total?", "Provide details for each pricing scheme.","What are the pricing schemes offered for [specific product/service]?","Show the pricing schemes sorted by [specific criterion, e.g., price, duration].","Display the pricing schemes applicable for [specific region/location].", "Are there any promotional pricing schemes available?","List the pricing schemes with discounts applied.","Show pricing schemes with tiered pricing structures.", "Can you display the pricing schemes with subscription options?","Provide details for the latest pricing schemes added.", "Show pricing schemes effective from [specific date or time period]."],
        "12": ["What is the pricing discount across different pricing schemes?","How does the pricing discount vary across various pricing methods?", "Can you provide a breakdown of pricing discounts under different pricing schemes?","What are the discount factors applied within each pricing method?", "How do the discount factors compare between different pricing schemes?","Which pricing scheme offers the highest discount, and by how much?", "Are there any trends in discount rates across different pricing methods?","How does the factor method influence pricing discounts compared to other methods?", "Can you analyze the impact of the factor method on overall pricing strategies?","Is there any correlation between the factor method and pricing discounts across schemes?", "How do pricing discounts under factor method compare with traditional pricing approaches?","Are there any anomalies in pricing discounts when using the factor method?", "Can you identify any outliers in pricing discounts within each pricing scheme?","What is the average pricing discount across all pricing schemes, considering the factor method?", "Are there any instances where the factor method leads to higher pricing discounts compared to other methods?"],
        "13": ["Print counts of Type 1 & 2 CPBs in the dataset.","Summarize the number of Type 1 & 2 CPBs available.","What is the count of Type 1 & 2 CPBs?", "Could you provide a summary of the quantity for Type 1 & 2 CPBs?","How many Type 1 & 2 CPBs are there in total?", "How many different types of products are available in the template?", "How many Type 1 in the template?", "How many Type 2 in the template?", "How many NSE in the template?", "Count of T1, T2, NSE?"],
        "14": ["NSE applicable to which products?", "Which all products NSE is applicable to?", "What is my scope of NSE? "]
    }
    
    return question_by_group

def build_group_context():
    group_ctx = ''
    
    for group_number, questions in question_by_group().items():
        group_ctx += f"'[GROUP_{group_number}]: {questions}' \n"  # Add group number and a newline
        # for question in questions:
        #     group_ctx += f"- {question}\n"  # Add each question with a hyphen and newline
    
    return group_ctx

# Define input question
def detectQueryGroup(user_input):
    if user_input:
        g_ctx = build_group_context()
        
        text=('Given the question: "'+user_input+'"?, and the following question groups: '
            + g_ctx +
            'Ask ChatGPT to determine which group the question exactly belongs to. If a result is detected for question groups, request a response from OpenAI in this format: "[GROUP_X]" or "NO_DATA". If no data is found, respond with "NO_DATA".'
            )
        
        response = client(text)
        return check_match(response)

def summarize(question, answer):
    context = f"Giving I have this question: {question} and this answer: '{str(answer)}'. Summarize the answer to make it match with the question in the human language and return the answer only."
    # print("Context: ", context)
    return client(context, role="user")
    # return agent.chat(context)

def filter_row_value_by_field(sheet_name, field_val, val_col_idx, field_idx=0):
    df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
    # Filter the DataFrame to get the row with field
    filtered_df = df[df[field_idx] == field_val]
    # Check for NaN and handle accordingly (assuming there's only one row)
    value = filtered_df.iloc[0, val_col_idx]
    if pd.isna(value):
        value = None  # Replace with None if NaN
        
    return value
    
def handle_default(question):
    return "I'm sorry, I don't understand the question."
    # return agent.chat(question)
    
def handle_response_G1(question):   
    sheet_name = 'Header'
    field = 'Case Type'
    value = filter_row_value_by_field(sheet_name, field, 1)

    return summarize(question, value)
    
def handle_response_G2(question):
    sheet_name = 'Header'
    field = 'Upload Type'
    value = filter_row_value_by_field(sheet_name, field, 1)

    return summarize(question, value)
    
def handle_response_G3(question):
    sheet_name = 'Header'
    field = 'End Customer'
    value = filter_row_value_by_field(sheet_name, field, 1)

    return summarize(question, value)

def handle_response_G4(question):
    sheet_name = 'Contextual data'
    field = 'Location Name'
    
    df = pd.read_excel(file_path, sheet_name=sheet_name, skiprows=3)
    # Filter location name
    filtered_data = df[field].dropna()
    value = filtered_data.values[0]
    if pd.isna(value):
        value = None  # Replace with None if NaN

    return summarize(question, value)

def handle_response_G5(question):
    sheet_name = 'Contextual data'
    
    df = pd.read_excel(file_path, sheet_name=sheet_name, skiprows=3)
    # Select columns up to the first NaN column
    selected_columns = df.columns[:df.columns.get_loc(df.isna().all().idxmax())]

    # Create a new dataframe with filtered columns
    phase_df = df[selected_columns].dropna()
    agent = Agent(phase_df, CONFIG)
    
    return agent.chat(question + ". And summary in human response.")

def handle_response_G6(question):
    sheet_name = 'Contextual data'
    
    df = pd.read_excel(file_path, sheet_name=sheet_name, skiprows=3)
    # Select columns up to the first NaN column
    selected_columns = df.columns[:df.columns.get_loc(df.isna().all().idxmax())]

    # Create a new dataframe with filtered columns
    phase_df = df[selected_columns].dropna()
    agent = Agent(phase_df, CONFIG)
    
    return agent.chat(question + ". And summary in human response.")

# How many tabs/worksheets are there in the file and how are they related?
def handle_response_G7(question):
    workbook = load_workbook(filename=file_path, read_only=True)

    # Get all sheet names
    all_sheet_names = workbook.sheetnames

    # Filter out hidden sheets
    visible_sheet_names = [sheet_name for sheet_name in all_sheet_names if not workbook[sheet_name].sheet_state == 'hidden']

    answer = f"{len(visible_sheet_names)} visible sheets: {', '.join(visible_sheet_names)}"

    return summarize(question=question, answer=answer)

# How many currencies the offer/contract transacted, what is the primary currency and what the exchange rate?
def handle_response_G8(question):
    contextual_data_sheet_name = "Contextual data"
    currency_df_headers = ['Currency', 'Exchange rate to', 'EUR']
    boq_sheet_name = "BoQ"
    response = None
    try: 
        # extract dataframe that contains Primary currency
        boq_df = pd.read_excel(file_path, sheet_name=boq_sheet_name, header=None, skiprows=30)
        field = "Primary Currency"
        primary_currency_row_index = boq_df[0] == field
        # Create a new DataFrame
        boq_primary_currency_df = pd.DataFrame({
            field: boq_df.loc[primary_currency_row_index, 1],
        })     
        print(f"boq data: {boq_primary_currency_df.to_string()}")
        
        # extract currency dataframe
        data_frame = get_sub_df(contextual_data_sheet_name, currency_df_headers)
        data_frame = data_frame.dropna()
        print(f"currency data: {data_frame.to_string()}")

        # Describe Dataframe Schema
        field_descriptions = {
            "Currency": "Name of the currency being exchanged.",
            "EUR": "Actual exchange rate."
        }

        # connector = PandasConnector({"original_df": data_frame}, field_descriptions=field_descriptions)
        # sdf = SmartDataframe(connector, config=CONFIG)
        agent = Agent([data_frame, boq_primary_currency_df], CONFIG)
        response = agent.chat(question)
    except Exception:
        response = "Response was not provided due to an error"
        pass # Handle the error silently   
    
    return summarize(question, response)

# Are all the contextual data used with the product boq?
def handle_response_G9(question):
    contextual_data_sheet_name = 'Contextual data'
    contextual_data_df = pd.read_excel(file_path, sheet_name=contextual_data_sheet_name, header=None, skiprows=3)
    contextual_data = SmartDataframe(contextual_data_df, name=contextual_data_sheet_name, description="Brief description of contextual data about \
                                     phases, customer location management, offering tags, offering categories, incoterms, and currencies")

    sheet_name1 = "Missing SI info"
    headers1 = ['SI Code', 'SI Name', 'SI Type', 'Design Type', 'Classification L3', 'Profit Center', 'IRP (base) (EUR)']
    data_frame1 = get_sub_df(sheet_name1, headers1)
    missing_si_info = SmartDataframe(data_frame1, name=sheet_name1, description="Missing SI information")
    
    sheet_name2 = "Payment Terms"
    headers2 = ['Application Scope (Type)', 'Application Scope (Value)', 'Days', 'Due From', 'SI Code']
    data_frame2 = get_sub_df(sheet_name2, headers2)
    payment_terms = SmartDataframe(data_frame2, name=sheet_name2, description="Payment Terms")

    boq_df1 = get_boq_df1()
    boq_df2 = get_boq_df2()
    
    boq_sdf_1 = SmartDataframe(boq_df1, name="BoQ", description="BoQ data part 1")
    boq_sdf_2 = SmartDataframe(boq_df2, name="BoQ", description="BoQ data part 2")
    
    # Handle with agent
    agent = Agent([contextual_data, missing_si_info, payment_terms, boq_sdf_1, boq_sdf_2],
                  config=CONFIG,
                  description="You are a data analysis agent. Your main goal is to help non-technical users to analyze data",
                )
    
    # Handle with agent
    agent = Agent([data_frame1, data_frame2, boq_df1, boq_df2], CONFIG)
    # response = agent.chat("Get data was not used from both 'Missing SI info' and 'Payment Terms' in the BoQ dataframe")
    response = agent.chat(question)
    return response
    
    # # User want to skip?
    # print()
    # while True:
    #     print(f"Clarification: ")
    #     # Get Clarification Questions
    #     questions = agent.clarification_questions(question)
    #     for clarify_query in questions:
    #         print(f"{clarify_query}")
            
    #     skip_input= input("Input the clarification query to tell me more or Input 'Skip' to ignore: ")
        
    #     if skip_input.lower().strip() == 'skip':
    #         break
        
    #     agent.chat(skip_input)
    
    # return ""

# Are there any duplicates on any of the sheets?
def handle_response_G10(question):
    """
    Reads an Excel spreadsheet, checks for duplicates across all sheets,
    and prints details of the duplicates if found.

    Args:
        excel_file (str): Path to the Excel file.
    """
    
    # Limit sheet to load to reduce data loading made the error
    # Request too large for gpt-3.5-turbo in organization org-reP5h1as4XuJ60BrwNV1ja7N on tokens per min (TPM): Limit 60000, Requested 17303596.
    sheets_to_load = [
        # "Header",
        # "Contextual data",
        "BoQ",
        # "Pricing Schemes"
    ]
    
    # Check for duplicates across all sheets, considering all columns
    duplicated_value = {}
    num_duplicates = 0
    has_duplicates = False
    sheet_names = pd.ExcelFile(file_path).sheet_names
    for sheet_name in sheet_names:
        if sheet_name in sheets_to_load:
            df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
            
            # Count non-NaN values in each row, excluding the entire row if all NaN
            non_nan_counts = df.notna().sum(axis=1)
            non_nan_counts = non_nan_counts[non_nan_counts > 0]  # Exclude rows with all NaN

            # Find the row with the maximum non-NaN count
            row_with_max_columns = non_nan_counts.idxmax()

            # Set the row with maximum non-NaN columns as DataFrame columns
            df.columns = df.iloc[row_with_max_columns]
            df.columns.name = ''

            # For demo
            if sheet_name == 'BoQ':
                df = df[['Element Type', 'Parent', 'External Reference']]    
            
            # Drop the row used as headers (optional)
            # df = df.iloc[1:]  # Uncomment to remove the row used for headers

            duplicated_rows = df[~df.isna().all(axis=1)].duplicated()
            # Print sheet name and duplicate rows if any
            if duplicated_rows.any():
                # print(f"Duplicates found in sheet: {sheet_name}")
                has_duplicates = True
                duplicated_value[sheet_name] = df[df.duplicated(keep=False) & ~df.isna().all(axis=1)]
                duplicated_value[sheet_name].index += 1 # To show actual row in dataset
                # Count the number of duplicates
                num_duplicates += duplicated_rows.sum()

    if not has_duplicates:
      return "No duplicates found in the Excel file."
    
    # combined_df = pd.concat(duplicated_value, ignore_index=False)  # Combine all DataFrames
    combined_str = f"""
    {duplicated_value}
    """
    response = f"{has_duplicates}. There are {num_duplicates} duplicated rows through sheets {sheets_to_load}: \n {combined_str}"
    
    return summarize(f"{question} If there are any duplicates then show the list of duplicated rows.", response)

# Show all the pricing schemes?
def handle_response_G11(question):
    sheet_name = "Pricing Schemes"
    # headers = ["Pricing Scheme Name","Pricing Scheme Type","Pricing Waterfall Group","Application Scope (Type)","Application Scope (Value)","Phases","Business Rationale","Method of Calculation","Factor","Value","Expiration Date","Other Condition","X Value","Y Value","Only Applicable to this Contract"]
    # sub_data_frame = get_sub_df(sheet_name, headers)
    # smartDataframe = init_smart_data_frame(sub_data_frame, sheet_name)
    # query = 'List all values?'
    # response = smartDataframe.chat(query)
    # return summarize(question, response)
    
    df = pd.read_excel(file_path, sheet_name=sheet_name, skiprows=1)
    # Select columns
    selected_columns = df.columns # [:df.columns.get_loc(df.isna().all().idxmax())]

    # Create a new dataframe with filtered columns
    pricing_scheme_df = df[selected_columns] #.dropna(axis=1)
    # print(f"Pricing: ", pricing_scheme_df)
    agent = Agent([pricing_scheme_df], CONFIG)
    response = agent.chat(f"Pricing Scheme Name can be empty. {question}. And summary in human response.")
    return response

# What is the pricing discount across pricing schemes?
def handle_response_G12(question):
    sheet_name = "Pricing Schemes"
    # headers = ["Pricing Scheme Name","Pricing Scheme Type","Pricing Waterfall Group","Application Scope (Type)","Application Scope (Value)","Phases","Business Rationale","Method of Calculation","Factor","Value","Expiration Date","Other Condition","X Value","Y Value","Only Applicable to this Contract"]
    # sub_data_frame = get_sub_df(sheet_name, headers)
    # smartDataframe = init_smart_data_frame(sub_data_frame, sheet_name)
    # query = 'List all values?'
    # response = smartDataframe.chat(query)
    # return summarize(question, response)
    
    df = pd.read_excel(file_path, sheet_name=sheet_name, skiprows=1)
    # Select columns
    selected_columns = df.columns # [:df.columns.get_loc(df.isna().all().idxmax())]

    # Create a new dataframe with filtered columns
    pricing_scheme_df = df[selected_columns] #.dropna(axis=1)
    # print(f"Pricing: ", pricing_scheme_df)
    agent = Agent([pricing_scheme_df], CONFIG)
    response = agent.chat(f"Look for value, factor and method columns for the pricing discount. {question}. Summary the answer!")
    # return agent.chat(f"{question}. Summary the answer!")
    return response

# How many type 1 & type 2 CPBs are there in the dataset?
def handle_response_G13(question):
    return generic_context.chat(question, tables=['BoQ'])

# NSE applicable to which products? 
def handle_response_G14(question):
    prompt = question + f"\nPlease provide detailed answer."
    return generic_context.chat(question, tables=['BoQ', 'NSE attached scope'])

# Create a dictionary to map input values to functions
function_map = {
    "1": handle_response_G1,
    "2": handle_response_G2,
    "3": handle_response_G3,
    "4": handle_response_G4,
    "5": handle_response_G5,
    "6": handle_response_G6,
    "7": handle_response_G7,
    "8": handle_response_G8,
    "9": handle_response_G9,
    "10": handle_response_G10,
    "11": handle_response_G11,
    "12": handle_response_G12,
    "13": handle_response_G13,
    "14": handle_response_G14
}

# Define a function to handle user questions and return responses
def handle_question(queryGroupNumber, question, raise_exeception=False):
    result = None
    # Check the user input and call the associated function
    try: 
        if str(queryGroupNumber) in function_map:
            function = function_map[queryGroupNumber]
            result = function(question)
        elif raise_exeception:
            raise ValueError('Cannot answer your question with group.') 
        else: 
            result = handle_default(question)
    except Exception as ex:
        if raise_exeception:
            raise ValueError(f'Cannot answer your question with group.') 
        else: 
            result = "I'm sorry, I don't understand the question."
                
    return result

# For Group 13:
# Init BoQ sub data frame
boq_sheet = 'BoQ'
boq_headers = [
    'Element Type', 
    'Parent', 
    'External Reference',
    'Description',
    'CPB Additional Attributes',
    'Local/Foreign',
    'Pricing Type',
    'Duration',
    'Pricing Model',
    'Pricing Metric',
    'Pricing Metric Factor',
    'Classification L1',
    'Classification L2',
    'Ordering Conditions',
    'Order Preparation Attributes',
    'Finalized Conf',
    'CPB Tags / Categories Attributes',
    'Category L1',
    'Category L2',
    'Course type', 
    'Course Level',
    'Delivery Type',
    'Curriculum Title',
    'Column7',
    'Column8',
    'Column9',
    'Column10',
    'SI Attributes',
    'Product - Service (ID)',
    'Product - Service (Description)',
    'Model (ID)',
    'Model (Description)',
    'Release (ID)',
    'Release (Description)',
    'WHAT (ID)',
    'WHAT (Description)',
    'PCI (ID)',
    'PCI (Description)',
    'Configured Profit Center (ID)', 
    'Configured Profit Center (Description)',
    'Config ID',
    'Cost Phase 1',
    'Cost Phase 2',
    'Cost Phase 3',
    'Cost Phase 4',
    'Cost Phase 5',
    'Cost Phase 6',
    'Cost Phase 7',
    'Cost Phase 8',
    'Cost Phase 9',
    'Cost Phase 10',
    'Cost Phase 11',
    'Cost Phase 12',
    'Cost Phase 13',
    'Cost Phase 14',
    'Cost Phase 15',
    'CLP Phase 1',
    'CLP Phase 2',
    'CLP Phase 3',
    'CLP Phase 4',
    'CLP Phase 5',
    'CLP Phase 6',
    'CLP Phase 7',
    'CLP Phase 8',
    'CLP Phase 9',
    'CLP Phase 10',
    'CLP Phase 11',
    'CLP Phase 12',
    'CLP Phase 13',
    'CLP Phase 14',
    'CLP Phase 15',
    'CNP Phase 1',
    'CNP Phase 2',
    'CNP Phase 3',
    'CNP Phase 4',
    'CNP Phase 5',
    'CNP Phase 6',
    'CNP Phase 7',
    'CNP Phase 8',
    'CNP Phase 9',
    'CNP Phase 10',
    'CNP Phase 11',
    'CNP Phase 12',
    'CNP Phase 13',
    'CNP Phase 14',
    'CNP Phase 15',
    'Reference Quantities',
    # 'Quantities_1',
    # 'Quantities_2',
    # 'Quantities_3',
    # 'Quantities_4',
    # 'Quantities_5',
    # 'Quantities_6',
    # 'Quantities_7',
    # 'Quantities_8',
    # 'Quantities_9',
    # 'Quantities_10',
    # 'Quantities_11',
    # 'Quantities_12',
    # 'Quantities_13',
    # 'Quantities_14',
    # 'Quantities_15',
    # 'Quantities_16',
    # 'Quantities_17',
    # 'Quantities_18',
    # 'Quantities_19',
    # 'Quantities_20',
    # 'Quantities_21',
    # 'Quantities_22',
    # 'Quantities_23',
    # 'Quantities_24',
    # 'Quantities_25', 
    # 'Quantities_26',
    # 'Quantities_27',
    # 'Quantities_28',
    # 'Quantities_29',
    # 'Quantities_30',
    # 'Quantities_31', 
    # 'Quantities_32',
    # 'Quantities_33',
    # 'Quantities_34',
    # 'Quantities_35',
    # 'Quantities_36',
    # 'Quantities_37',
    # 'Quantities_38',
    # 'Quantities_39',
    # 'Quantities_40',
    # 'Quantities_41',
    # 'Quantities_42',
    # 'Quantities_43',
    # 'Quantities_44',
    # 'Quantities_45',
    # 'Quantities_46',
    # 'Quantities_47',
    # 'Quantities_48', 
    # 'Quantities_49', 
    # 'Quantities_50', 
    # 'Quantities_51',
    # 'Quantities_52',
    # 'Quantities_53'
]
# boq_df = get_sub_df(boq_sheet, boq_headers)
def get_boq_df(custom_header=boq_headers):
    if custom_header:
        return get_sub_df(boq_sheet, custom_header)
    else:
        return get_sub_df(boq_sheet, boq_headers)

boq_sheet1 = 'BoQ'
boq_headers1 = [
    'Element Type', 
    'Parent',
    'External Reference',
    'Description',
    'CPB Additional Attributes',
    'Local/Foreign',
    'Pricing Type',
    'Duration',
    'Pricing Model',
    'Pricing Metric',
    'Pricing Metric Factor',
    'Classification L1',
    'Classification L2',
    'Ordering Conditions',
    'Order Preparation Attributes',
    'Finalized Conf',
    'CPB Tags / Categories Attributes',
    'Category L1',
    'Category L2',
    'Course type', 
    'Course Level',
    'Delivery Type',
    'Curriculum Title',
    'Column7',
    'Column8',
    'Column9',
    'Column10',
    'SI Attributes',
    'Product - Service (ID)',
    'Product - Service (Description)',
    'Model (ID)',
    'Model (Description)',
    'Release (ID)',
    'Release (Description)',
    'WHAT (ID)',
    'WHAT (Description)',
    'PCI (ID)',
    'PCI (Description)',
    'Configured Profit Center (ID)', 
    'Configured Profit Center (Description)',
    'Config ID'
]

# boq_df1 = get_sub_df(boq_sheet1, boq_headers1)
def get_boq_df1():
    return get_sub_df(boq_sheet1, boq_headers1)

boq_sheet2 = 'BoQ'
boq_headers2 = [
    'Element Type', 
    'Parent', 
    
    'External Reference',
    'Cost Phase 1',
    'Cost Phase 2',
    'Cost Phase 3',
    'Cost Phase 4',
    'Cost Phase 5',
    'Cost Phase 6',
    'Cost Phase 7',
    'Cost Phase 8',
    'Cost Phase 9',
    'Cost Phase 10',
    'Cost Phase 11',
    'Cost Phase 12',
    'Cost Phase 13',
    'Cost Phase 14',
    'Cost Phase 15',
    'CLP Phase 1',
    'CLP Phase 2',
    'CLP Phase 3',
    'CLP Phase 4',
    'CLP Phase 5',
    'CLP Phase 6',
    'CLP Phase 7',
    'CLP Phase 8',
    'CLP Phase 9',
    'CLP Phase 10',
    'CLP Phase 11',
    'CLP Phase 12',
    'CLP Phase 13',
    'CLP Phase 14',
    'CLP Phase 15',
    'CNP Phase 1',
    'CNP Phase 2',
    'CNP Phase 3',
    'CNP Phase 4',
    'CNP Phase 5',
    'CNP Phase 6',
    'CNP Phase 7',
    'CNP Phase 8',
    'CNP Phase 9',
    'CNP Phase 10',
    'CNP Phase 11',
    'CNP Phase 12',
    'CNP Phase 13',
    'CNP Phase 14',
    'CNP Phase 15',
    'Reference Quantities'
]

# boq_df2 = get_sub_df(boq_sheet2, boq_headers2)
def get_boq_df2():
    return get_sub_df(boq_sheet2, boq_headers2)

def main():
    while True:
        user_input = input("Ask me a question about the data: ")
        if user_input.lower() == 'exit':
            break
        
        queryGrNumber = detectQueryGroup(user_input)
        print("For debug testing. Gr: ", queryGrNumber)
        response = handle_question(queryGrNumber, user_input)
        if response is not None and response:
            print("Assistant: ", response)
            print()

def chat(user_input, load_suggested_replies=False):
    response = {}
    queryGrNumber = detectQueryGroup(user_input)
    logger.info(f"For debug testing. Gr: {queryGrNumber}. Question: {user_input}")
    answer = handle_question(queryGrNumber, user_input)
    if not isinstance(answer, str):
        response["Assistant"] = str(answer)  # Convert non-strings using str()
    else:
        response["Assistant"] = answer

    if (load_suggested_replies):
        response["suggested_replies"] = suggested_replies(queryGrNumber, user_input)
    
    return response

# def chat_with_answer(user_input):
#     queryGrNumber = detectQueryGroup(user_input)
#     logger.info(f"For debug testing. Gr: {queryGrNumber}. Question: {user_input}")
#     answer = handle_question(queryGrNumber, user_input, raise_exeception=True)
    
#     return (queryGrNumber, answer)

def chat_with_answer(queryGrNumber, user_input):
    logger.info(f"For debug testing. Gr: {queryGrNumber}. Question: {user_input}")
    answer = handle_question(queryGrNumber, user_input, raise_exeception=True)
    
    return answer

def set_file_path(new_path):
    global file_path
    file_path = new_path
    # global data_frame_manager
    # data_frame_manager = DataFrameManager(new_path)

if __name__ == "__main__":
    main()
